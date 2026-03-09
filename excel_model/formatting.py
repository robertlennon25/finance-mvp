from __future__ import annotations

from openpyxl.styles import Font, PatternFill


SECTION_FILL = PatternFill(fill_type="solid", fgColor="D9E2F3")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")


def add_header(ws, title: str) -> None:
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)


def style_section(cell) -> None:
    cell.font = Font(bold=True)
    cell.fill = SECTION_FILL


def style_header_row(ws, row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL


def autosize_columns(ws) -> None:
    for col_cells in ws.columns:
        max_length = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 26)
